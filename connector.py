import requests
import re
import string
from openpyxl import Workbook
from openpyxl import load_workbook
import json
from bs4 import BeautifulSoup as bs
from yahoofinancials import YahooFinancials as yafin

URL = "https://finviz.com/quote.ashx?"
rts_codes = ['UBER','CTVA','DD','CCK','RAMP','EPC','ARW','MTH','CLGX','ATGE','AVNS','PEN','BLD','GCO','ENV','NGVT','FLOW',
            'BFAM','CPS','AXE','VRTU','VRTV','HGV','CRMT','APEI','CNXN','VPG','MYRG','SAIL','Y','SFIX','FORR','MCRI','ROKU','LITE',
            'MTCH','ZS','SNBR','COUP','BECN','IAC','CARG','IONS','MEDP','GCP','AAXN','UFPI','SIVB','EYE','AZPN','SYNH','TCBI','PINC',
            'COHR','ALRM','MASI','LGIH','PBH','TREE','OMCL','ICUI','ANAB','FIZZ','MYOK','FOXF','LHCG','VICR','QDEL','SYKE','SAIA',
            'PLXS','CENTA','VREX','ALLK','NEOG','SPSC','ITRI','WGO','CENT','ROCK','BAND','NANO','DORM','PETQ','CSWI','OFIX','HCCI',
            'HURN','ROLL','MSTR','CRVL','CVCO','SRDX','SP','FRPH','LEVI','LYFT','TWLO','ASIX','TPIC','RYTM','TTD','DISCB','KTB','DOW',
            'QTNA','W','DK','BURL','GRUB','MEI','TDS','CHE','WOR','IBP','CTB','DRQ','WHD','AYX','BKI','SERV','BID','SSTK','KFY','VMW',
            'PBF','ARMK','SCCO','KNX','KMT','WCC','CNK','PFGC','DLB','YELP','ZEN','WWE','GDDY','VNE','RYN','HAE','IQV','ELAN','PLNT',
            'CLH','WEX','SITE','BH','AMN','CRL','FND','WCG','POL','GRA','DKS','RH','MATX','TOL','TPX','TKR','WWW','MD','VC','MLHR',
            'IIVI','TER','EPAY','ANDE','SBGI','TWOU','ABMD','ATRO','PATK','STLD','CMCO','SCSC','SEDG','HA','IRBT','ANIK','PRGS','BRKR',
            'STRA','NUVA','RAVN','TRMB','WDAY','BABY','NTGR','CY','MYGN','DIOD','MINI','ON','FIVE','CHDN','PZZA','TTWO','WERN','ALGT',
            'RRGB','CRUS','LOGM','WING','FOX','FOXA','DIS','BEB','KCEL','KZTK','LTHM','TNET','TREX','UFS','UNF','USM','WBC','WTS','SSD',
            'SXI','SXT','TDY','TRU','SMG','AAN','ABG','AIR','ALV','ASGN','BCO','BDC','BERY','AWI','AWR','BMI','CBM','CE','DLX','EBS','FCN',
            'FUL','GBX','GDOT','GEF','GHC','GMED','GNRC','GPI','CR','CRS','GVA','HXL','KEX','LAD','MOV','MSM','MTRN','NEU','NJR','NOW',
            'NSP','NUS','MMS','PRLB','RGR','ROG','ROL','SAM','SAVE','REX','RGEN','SINA','PRSC','MMSI','NXST','OSIS','PCTY','PEGA','POWI',
            'NSIT','MTSC','MANH','MANT','MGLN','IDCC','IPAR','JBSS','JOUT','KALU','HUBG','CVGW','GTLS','ENSG','ENTA','ERIE','EXLS','FARO',
            'CGNX','CPRT','CALM','BBSI','BJRI','ATRI','AMWD','AEIS','SMTC','LOPE','ACIA','QADA','THRM','WWD','WAB','AIV','ECA','CVET','AGCO',
            'ALSN','ANET','AOS','ASH','BAH','BC','BIG','BIO','BR','CRI','CSL','DDS','DECK','DY','EPAM','CLR','EXP','GWR','GWRE','H','HFC',
            'HII','INGR','JLL','KEYS','LEA','FDS','FLT','LII','LVS','MAN','MSCI','MTN','OSK','PAYC','PII','PKG','NVR','OC','RPM','RS','RMD',
            'SHAK','SNX','SPR','TYL','VEEV','WLK','WSM','WSO','XPO','TFX','THO','TXRH','ZBRA','UBNT','UTHR','SSNC','TECD','SLAB','RP','SAFM',
            'SEIC','ODFL','OLED','PLAY','PLCE','POOL','PRAH','QLYS','MXIM','NDSN','MDSO','MELI','MIDD','MKTX','LULU','FTNT','LECO','IPGP','JCOM',
            'JKHY','HQY','IART','INGN','HCSG','FANG','COLM','ETSY','CASY','CBRL','CDNS','BLKB','AVAV','AMCX','AMED','CI','D','ETRN','WRK','LIN',
            'REZI','ET','GTX','PRSP','WH','APY','SPLK','AVGO','PRIK01','ARNC','DLPH','SQ','PANW','LEN.B','ETM','BMW','CBPO','DWDP','RJF','DLR',
            'ALGN','ALK','GPN','LLL','LKQ','FTI','MAA','TMK','SNPS','JCI','SPGI','BHF','ARE','HBAN','FL','AJG','TDG','ALB','ULTA','IDXX','FBHS',
            'AEE','REG','RE','CHTR','VNO','COTY','LNT','CNC','ED','AYI','COO','INCY','MTD','UA','GS','AMD','INFO','WLTW','FTR','IT','HOLX','DXC',
            'HLT','FTV','ANSS','UNM','BHGE','XRX','AA','URBN','HBI','AZO','UAA','STZ','XRAY','SIG','KMX','NWL','APTV','XLNX','VFC','MPC','NTRS',
            'WDC','XEC','ALXN','USB','STT','DRI','MMC','BF.B','ADS','CMS','HST','PBI','IR','FLIR','HUM','PBCT','RTN','EQIX','LEG','GPC','EXPE',
            'KIM','BKNG','JWN','RF','ENDP','M','NEE','SYF','AMAT','CAH','WU','KDP','MUR','KHC','SYMC','VRTX','BWA','TROW','GPS','CRM','TRIP','BXP',
            'DISCA','NOV','HP','ANTM','HSIC','CTSH','APD','COST','SO','TDC','RCL','NWS','RL','ESS','DTE','VTR','PRU','PNC','TXN','DISCK','WY','CB',
            'CMG','SWKS','AMG','CFG','MCO','OKE','IRM','EA','NWSA','ZION','FISV','OI','ECL','KR','FFIV','QRVO','MNST','IFF','HIG','SLG','BDX','SEE',
            'MYL','AKAM','VRSN','IPG','SHW','K','GRMN','BBY','DVA','PEG','CMCSA','CHD','MOS','ETR','ISRG','RRC','RHT','PLD','ADI','GLW','REGN','AVY',
            'FIS','LM','WAT','VIAB','HPE','ORLY','ES','AAP','WELL','HOG','COF','ICE','MCHP','CERN','BEN','IP','AES','WHR','APH','LOW','JNPR','BK','CTL',
            'WEC','L','MAR','MKC','CF','DHI','HRB','ADBE','DFS','CL','SWN','CINF','SJM','TJX','STX','KEY','ROST','MSI','PSA','PHM','AMT','LLY','EQT',
            'HCP','WYNN','IVZ','FE','HPQ','TSS','FITB','WBA','DGX','LYB','ZBH','HRS','AIZ','NDAQ','A','BRK.B','MHK','NBL','HAS','BAX','CTXS','EW','HRL',
            'TGNA','FMC','PGR','LB','PCG','SYK','PVH','ADSK','ILMN','HCA','TSCO','JEF','ABT','AAL','RIG','CAG','VAR','TRV','CBRE','CCI','BBBY','AFL','PKI',
            'INTU','SRE','PFG','UHS','AMP','CVS','KMB','ATVI','CLX','ALLE','ADM','EL','UDR','AON','CPB','TEL','EXR','WYND','NTAP','BBT','CCL','ZTS','MAC',
            'OMC','NUE','YUM','TMO','GT','BLL','AN','TSN','VMC','TPR','PPL','COG','EMN','KLAC','ADP','MCK','TAP','MAT','MTB','AVB','CMA','MLM','PDCO','NAVI',
            'LH','SYY','MNK','O','GIS','CPRI','SPG','XEL','EIX','LEN','PPG','LRCX','CNP','ALL','DG','CMI','PCAR','PEP','BMY','DHR','APC','SLB','BLK','FLR',
            'MO','MDT','HON','UNH','TGT','AME','FAST','ETN','DE','PH','NVDA','ROP','NKE','C','LMT','HES','FDX','HD','NLSN','UNP','CELG','AGN','OXY','DLTR',
            'LUV','CHRW','ITW','WM','AMGN','RHI','DVN','MMM','KMI','UTX','NSC','SNA','MA','GWW','PXD','TXT','UAL','HAL','UPS','DOV','BIIB','SWK','URI',
            'FLS','EXPD','AXP','PSX','ROK','JEC','KSU','PWR','VRSK','EFX','R','MAS','MDLZ','WFC','WMB','FCX','MRO','CXO','RSG','JPM','CTAS','COP','SRCL',
            'ACN','SCHW','JBHT','MRK','HSY','NOC','APA','EOG','XYL','GM','EMR','AIG','ORCL','BSX','GD','GDWS','BAST','AKZM','CLF','KTF','RACE','GOOG',
            'GOOGL','PYPL','TIF','IRAO','SPB','DAL','AVP','MSFT','SBUX','FB','EBAY','NEM','CSCO','GILD','IBM','EXC','CBS','FSLR','JNJ','QCOM','CME','PFE',
            'MCD','WMT','VLO','BA','MS','T','PM','BAC','INTC','VZ','AMZN','CHK','CAT','GE','PG','ETFC','TSLA','AAPL','XOM','TWTR','NFLX','V','AABA','KO',
            'ABBV','NRG','F','MU','MET','CVX']

def request_page(url: str, ticker: str = "AAPL"):
    """requesting html page from finviz using default AAPL """
    try:
        payload = {"t": ticker}
        req = requests.get(url, params=payload)
        data = bs(req.content, 'html.parser')
        if req.status_code == 200:
            data = bs(req.content, 'lxml')
            return data
        else:
            print ("couldn't get page, status code: ",req.status_code) 
            return
    except Exception as e:
        print (e)
        return
        

def find_data(company: str, multiplicators: list):
    """is looking for company's (like 'FB'/'AAPL') parametrs like ROE, P/E in html page """
    data = request_page(URL, company)
    companys_multi = {}
    try:
        print (data.title.text)
        for param in multiplicators:
            result = data.find(text = param)
            multi = result.next_element.text
            companys_multi[result] = multi
        print (companys_multi)
        return companys_multi
    except Exception as e:
        print (e)


def get_urls(*params: str, num_of_links: int = 1) -> list:
    """because of finviz's limitations, one url can not be used
    Thats why ulrs are splitted if urls length is more than 2900
    fa_pe_u50      p/e<50
    fa_pe_o10      p/e>10

    fa_roe_u00   roe<00
    fa_roe_o15   roe>15

    fa_debteq_u15   debt/equity<15
    fa_debteq_o10   debt/equity>10

    fa_roi_o20    roi>20
    fa_roi_u10    roi<10

    fa_ps_u5    p/s<5
    fa_ps_o1    P/S>1"""
    urls = []
    try:
        for i in range(num_of_links):
            url = "https://finviz.com/screener.ashx?v=111"
            codes = ','.join(rts_codes[len(rts_codes)*(num_of_links - i - 1)//num_of_links:(len(rts_codes)*(num_of_links - i)//num_of_links)])
            payload = {"FT": 2,"f": params,"t": codes}
            req = requests.get(url, params=payload)
            if len(req.url) > 2900:
                urls = []
                num_of_links += 1
                urls = get_urls(*params, num_of_links=num_of_links)
            else:
                urls.append(req.url)
        return (urls)
    except Exception as e:
        print (e)
        return None
    

def get_balance(ticker):
    all_balance = yafin(ticker)
    print (json.dumps(all_balance.get_financial_stmts("annual", "income"), sort_keys=True, indent=4))
    print ((all_balance.get_operating_income()))

def get_macro(company, docs):
#    while True:
#        docs = int(input("Choose:\n 1.Income-statement\n 2.Balance-sheet\n"))
#        if docs == 1:
#            docs = "income-statement"
#            break
#        if docs == 2:
#            docs = "balance-sheet"
#            break
    req = requests.get("https://www.macrotrends.net/stocks/charts/"+company+"/"+docs)
    data = bs(req.content, 'html.parser')
    result = data.find_all('script')
    var = (re.findall(r"^ var originalData = (.+);", str(result), re.M))
    listik = json.loads(var[0])
    last_dict = {}
    for dictionary in listik:
        del dictionary["popup_icon"]
        changed_field_name = re.findall(r">(.+)</a>", dictionary["field_name"])
        dictionary["field_name"] = changed_field_name
        if dictionary.get("field_name") != []:
            for key, value in sorted(dictionary.items()):
                if key != "field_name":
                    if not last_dict.get(key):
                        last_dict[key] = {}
                        field_name = dictionary["field_name"][0]
                        last_dict[key][field_name] = value
                    else:
                        field_name = dictionary["field_name"][0]
                        last_dict[key][field_name] = value
    return last_dict

def write_to_exel(our_dict, sheet):
    book = load_workbook("test.xlsx")
    ws = book.create_sheet(sheet)
    i, j = 2, 2
    for key in sorted(our_dict.keys()):
        ws.cell(row = 1, column = j).value = key
        for key1, value1 in sorted(our_dict[key].items()):
            if ws.cell(row=i, column=1).value != key1:
                ws.cell(row=i, column=1).value = key1
                ws.cell(row=i, column=j).value = value1
                i += 1
            else:
                ws.cell(row=i, column=j).value = value1
                i += 1
        i = 2
        j += 1
    book.save("test.xlsx")

if __name__ == "__main__":
#    my_dict = get_macro("https://www.macrotrends.net/stocks/charts/AAPL/apple/income-statement")
    my_dict1 = get_macro("AAPL/apple", "income-statement")
    my_dict2 = get_macro("AAPL/apple", "balance-sheet")
    write_to_exel(my_dict1, "income-statement")
    write_to_exel(my_dict2, "balance-sheet")
#    for key,value in sorted(last_dict.items()):
#        print (key, value)

#    get_balance("AAPL")
#    screener("fa_pe_u15")
#    print (get_urls("fa_pe_u15", "fa_debteq_u0.1","fa_roe_o25"))
#    find_data(company="FB", multiplicators=["P/E","P/B","ROE","ROA","P/S"])
