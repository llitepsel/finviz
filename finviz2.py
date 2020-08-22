import requests
import sys
import re
import string
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
import json
from bs4 import BeautifulSoup as bs
from yahoofinancials import YahooFinancials as yafin
import yfinance as yf

URL = "https://finviz.com/quote.ashx?"
"""
rts_codes = ['UBER','CTVA','DD','CCK','RAMP','EPC','ARW','MTH','CLGX','ATGE','AVNS','PEN','BLD','GCO','ENV','NGVT','FLOW',
            'BFAM','CPS','AXE','VRTU','VRTV','HGV','CRMT','APEI','CNXN','VPG','MYRG','SAIL','Y','SFIX','FORR','MCRI','ROKU','LITE',
            'MTCH','ZS','SNBR','COUP','BECN','IAC','CARG','IONS','MEDP','GCP','AAXN','UFPI','SIVB','EYE','AZPN','SYNH','TCBI','PINC',
            'COHR','ALRM','MASI','LGIH','PBH','TREE','OMCL','ICUI','ANAB','FIZZ','MYOK','FOXF','LHCG','VICR','QDEL','SYKE','SAIA',
            'PLXS','CENTA','VREX','ALLK','NEOG','SPSC','ITRI','WGO','CENT','ROCK','BAND','NANO','DORM','PETQ','CSWI','OFIX','HCCI',
            'HURN','ROLL','MSTR','CRVL','CVCO','SRDX','SP','FRPH','LEVI','LYFT','TWLO','ASIX','TPIC','RYTM','TTD','DISCB','KTB','DOW',
            'QTNA','W','DK','BURL','GRUB','MEI','TDS','CHE','WOR','IBP','CTB','DRQ','WHD','AYX','BKI','SERV','BID','SSTK','KFY','VMW',
            'PBF','ARMK','SCCO','KNX','KMT','WCC','CNK','PFGC','DLB','YELP','ZEN','WWE','GDDY','VNE','RYN','HAE','IQV','ELAN','PLNT',
            'CLH','WEX','SITE','BH','AMN','CRL','FND','WCG','POL','GRA','DKS','RH','MATX','TOL','TPX','TKR','WWW','MD','VC','MLHR',
            'IIVI','TER','EPAY','ANDE','SBGI','TWOU','ABMD','ATRO','PATK','STLD','CMCO','SCSC','SEDG','HA','IRBT','ANIK','PRGS','BRKR',
            'STRA','NUVA','RAVN','TRMB','WDAY','BABY','NTGR','CY','MYGN','DIOD','MINI','ON','FIVE','CHDN','PZZA','TTWO','WERN','ALGT',
            'RRGB','CRUS','LOGM','WING','FOX','FOXA','DIS','BEB','KCEL','KZTK','LTHM','TNET','TREX','UFS','UNF','USM','WBC','WTS','SSD',
            'SXI','SXT','TDY','TRU','SMG','AAN','ABG','AIR','ALV','ASGN','BCO','BDC','BERY','AWI','AWR','BMI','CBM','CE','DLX','EBS','FCN',
            'FUL','GBX','GDOT','GEF','GHC','GMED','GNRC','GPI','CR','CRS','GVA','HXL','KEX','LAD','MOV','MSM','MTRN','NEU','NJR','NOW',
            'NSP','NUS','MMS','PRLB','RGR','ROG','ROL','SAM','SAVE','REX','RGEN','SINA','PRSC','MMSI','NXST','OSIS','PCTY','PEGA','POWI',
            'NSIT','MTSC','MANH','MANT','MGLN','IDCC','IPAR','JBSS','JOUT','KALU','HUBG','CVGW','GTLS','ENSG','ENTA','ERIE','EXLS','FARO',
            'CGNX','CPRT','CALM','BBSI','BJRI','ATRI','AMWD','AEIS','SMTC','LOPE','ACIA','QADA','THRM','WWD','WAB','AIV','ECA','CVET','AGCO',
            'ALSN','ANET','AOS','ASH','BAH','BC','BIG','BIO','BR','CRI','CSL','DDS','DECK','DY','EPAM','CLR','EXP','GWR','GWRE','H','HFC',
            'HII','INGR','JLL','KEYS','LEA','FDS','FLT','LII','LVS','MAN','MSCI','MTN','OSK','PAYC','PII','PKG','NVR','OC','RPM','RS','RMD',
            'SHAK','SNX','SPR','TYL','VEEV','WLK','WSM','WSO','XPO','TFX','THO','TXRH','ZBRA','UBNT','UTHR','SSNC','TECD','SLAB','RP','SAFM',
            'SEIC','ODFL','OLED','PLAY','PLCE','POOL','PRAH','QLYS','MXIM','NDSN','MDSO','MELI','MIDD','MKTX','LULU','FTNT','LECO','IPGP','JCOM',
            'JKHY','HQY','IART','INGN','HCSG','FANG','COLM','ETSY','CASY','CBRL','CDNS','BLKB','AVAV','AMCX','AMED','CI','D','ETRN','WRK','LIN',
            'REZI','ET','GTX','PRSP','WH','APY','SPLK','AVGO','PRIK01','ARNC','DLPH','SQ','PANW','LEN.B','ETM','BMW','CBPO','DWDP','RJF','DLR',
            'ALGN','ALK','GPN','LLL','LKQ','FTI','MAA','TMK','SNPS','JCI','SPGI','BHF','ARE','HBAN','FL','AJG','TDG','ALB','ULTA','IDXX','FBHS',
            'AEE','REG','RE','CHTR','VNO','COTY','LNT','CNC','ED','AYI','COO','INCY','MTD','UA','GS','AMD','INFO','WLTW','FTR','IT','HOLX','DXC',
            'HLT','FTV','ANSS','UNM','BHGE','XRX','AA','URBN','HBI','AZO','UAA','STZ','XRAY','SIG','KMX','NWL','APTV','XLNX','VFC','MPC','NTRS',
            'WDC','XEC','ALXN','USB','STT','DRI','MMC','BF.B','ADS','CMS','HST','PBI','IR','FLIR','HUM','PBCT','RTN','EQIX','LEG','GPC','EXPE',
            'KIM','BKNG','JWN','RF','ENDP','M','NEE','SYF','AMAT','CAH','WU','KDP','MUR','KHC','SYMC','VRTX','BWA','TROW','GPS','CRM','TRIP','BXP',
            'DISCA','NOV','HP','ANTM','HSIC','CTSH','APD','COST','SO','TDC','RCL','NWS','RL','ESS','DTE','VTR','PRU','PNC','TXN','DISCK','WY','CB',
            'CMG','SWKS','AMG','CFG','MCO','OKE','IRM','EA','NWSA','ZION','FISV','OI','ECL','KR','FFIV','QRVO','MNST','IFF','HIG','SLG','BDX','SEE',
            'MYL','AKAM','VRSN','IPG','SHW','K','GRMN','BBY','DVA','PEG','CMCSA','CHD','MOS','ETR','ISRG','RRC','RHT','PLD','ADI','GLW','REGN','AVY',
            'FIS','LM','WAT','VIAB','HPE','ORLY','ES','AAP','WELL','HOG','COF','ICE','MCHP','CERN','BEN','IP','AES','WHR','APH','LOW','JNPR','BK','CTL',
            'WEC','L','MAR','MKC','CF','DHI','HRB','ADBE','DFS','CL','SWN','CINF','SJM','TJX','STX','KEY','ROST','MSI','PSA','PHM','AMT','LLY','EQT',
            'HCP','WYNN','IVZ','FE','HPQ','TSS','FITB','WBA','DGX','LYB','ZBH','HRS','AIZ','NDAQ','A','BRK.B','MHK','NBL','HAS','BAX','CTXS','EW','HRL',
            'TGNA','FMC','PGR','LB','PCG','SYK','PVH','ADSK','ILMN','HCA','TSCO','JEF','ABT','AAL','RIG','CAG','VAR','TRV','CBRE','CCI','BBBY','AFL','PKI',
            'INTU','SRE','PFG','UHS','AMP','CVS','KMB','ATVI','CLX','ALLE','ADM','EL','UDR','AON','CPB','TEL','EXR','WYND','NTAP','BBT','CCL','ZTS','MAC',
            'OMC','NUE','YUM','TMO','GT','BLL','AN','TSN','VMC','TPR','PPL','COG','EMN','KLAC','ADP','MCK','TAP','MAT','MTB','AVB','CMA','MLM','PDCO','NAVI',
            'LH','SYY','MNK','O','GIS','CPRI','SPG','XEL','EIX','LEN','PPG','LRCX','CNP','ALL','DG','CMI','PCAR','PEP','BMY','DHR','APC','SLB','BLK','FLR',
            'MO','MDT','HON','UNH','TGT','AME','FAST','ETN','DE','PH','NVDA','ROP','NKE','C','LMT','HES','FDX','HD','NLSN','UNP','CELG','AGN','OXY','DLTR',
            'LUV','CHRW','ITW','WM','AMGN','RHI','DVN','MMM','KMI','UTX','NSC','SNA','MA','GWW','PXD','TXT','UAL','HAL','UPS','DOV','BIIB','SWK','URI',
            'FLS','EXPD','AXP','PSX','ROK','JEC','KSU','PWR','VRSK','EFX','R','MAS','MDLZ','WFC','WMB','FCX','MRO','CXO','RSG','JPM','CTAS','COP','SRCL',
            'ACN','SCHW','JBHT','MRK','HSY','NOC','APA','EOG','XYL','GM','EMR','AIG','ORCL','BSX','GD','GDWS','BAST','AKZM','CLF','KTF','RACE','GOOG',
            'GOOGL','PYPL','TIF','IRAO','SPB','DAL','AVP','MSFT','SBUX','FB','EBAY','NEM','CSCO','GILD','IBM','EXC','CBS','FSLR','JNJ','QCOM','CME','PFE',
            'MCD','WMT','VLO','BA','MS','T','PM','BAC','INTC','VZ','AMZN','CHK','CAT','GE','PG','ETFC','TSLA','AAPL','XOM','TWTR','NFLX','V','AABA','KO',
            'ABBV','NRG','F','MU','MET','CVX']
"""
rts_codes = ['A','AA','AAL','AAN','AAP','AAPL','AAXN','ABBV','ABC','ABG','ABMD','ABT','ACAD','ACH','ACIA','ACM',
		'ACN','ADBE','ADI','ADM','ADP','ADS','ADSK','ADUS','AEE','AEIS','AERI','AES','AFG','AFL','AGCO','AGIO',
		'AIG','AIMT','AIR','AIV','AIZ','AJG','AJRD','AKAM','AKZM','ALB','ALGN','ALGT','ALK','ALL','ALLE','ALLK',
		'ALLO','ALNY','ALRM','ALSN','ALTR','ALV','ALXN','AMAT','AMCX','AMD','AME','AMED','AMG','AMGN','AMN','AMP',
		'AMT','AMWD','AMZN','AN','ANAB','ANDE','ANET','ANGI','ANIK','ANIP','ANSS','ANTM','AON','AOS','APA','APD',
		'APEI','APH','APLE','APPF','APPN','APTV','ARE','ARMK','ARNA','ARNC','ARW','ARWR','ASGN','ASH','ASIX','ATGE',
		'ATKR','ATR','ATRA','ATRI','ATRO','ATUS','ATVI','AVAV','AVB','AVGO','AVLR','AVNS','AVY','AWI','AWK','AWR',
		'AX','AXE','AXGN','AXP','AXSM','AYI','AYX','AZO','AZPN','BA','BABA','BAC','BAH','BAND','BAST','BAX','BBBY',
		'BBSI','BBY','BC','BCO','BCPC','BDC','BDX','BEAT','BECN','BEN','BERY','BF.B','BFAM','BFYT','BH','BHF','BIDU',
		'BIG','BIIB','BILI','BIO','BJRI','BK','BKI','BKNG','BKR','BL','BLD','BLDR','BLK','BLKB','BLL','BLUE','BMCH',
		'BMI','BMRN','BMW','BMY','BOH','BOKF','BOOT','BPMC','BR','BRC','BRK.B','BRKR','BRO','BSX','BTI','BUD','BURL',
		'BWA','BWXT','BXP','BYND','BZUN','C','CABO','CACC','CAG','CAH','CALM','CARA','CARG','CARR','CARS','CASY','CAT',
		'CB','CBPO','CBRE','CBRL','CBSH','CBU','CCI','CCK','CCL','CCMP','CDK','CDNA','CDNS','CDW','CE','CEA','CENT','CENTA',
		'CERN','CF','CFG','CFR','CFX','CGNX','CHA','CHD','CHDN','CHE','CHEF','CHGG','CHH','CHK','CHL','CHNG','CHRW','CHTR',
		'CHX','CI','CIEN','CINF','CL','CLDT','CLF','CLGX','CLH','CLR','CLX','CMA','CMCO','CMCSA','CME','CMG','CMI','CMP',
		'CMS','CNC','CNK','CNP','CNXN','COF','COG','COHR','COLM','COO','COP','CORR','CORT','COST','COTY','COUP','CPB','CPRI',
		'CPRT','CPS','CR','CREE','CRI','CRL','CRM','CRMT','CROX','CRS','CRUS','CRVL','CSCO','CSGP','CSII','CSL','CSOD','CSWI',
		'CSX','CTAS','CTB','CTL','CTLT','CTSH','CTVA','CTXS','CVCO','CVET','CVGW','CVLT','CVS','CVX','CW','CXO','D','DAL','DAR',
		'DBX','DCI','DD','DDOG','DDS','DE','DECK','DFS','DG','DGX','DHI','DHR','DIOD','DIS','DISCA','DISCB','DISCK','DK','DKS',
		'DLB','DLPH','DLR','DLTH','DLTR','DLX','DNLI','DNOW','DOCU','DORM','DOV','DOW','DPZ','DRI','DRQ','DVA','DVN','DXC','DXCM',
		'DY','EA','EBAY','EBS','ECHO','ECL','ECPG','ED','EDIT','EEFT','EFX','EGHT','EGRX','EHTH','EIX','EL','ELAN','EME','EMN',
		'EMR','ENDP','ENS','ENSG','ENTA','ENTG','ENV','EOG','EPAM','EPAY','EPC','EQIX','EQT','ERIE','ES','ESPR','ESS','ET','ETFC',
		'ETN','ETR','ETRN','ETSY','EV','EVBG','EVH','EVR','EW','EWBC','EXAS','EXC','EXEL','EXLS','EXP','EXPD','EXPE','EXR','EYE',
		'F','FANG','FARO','FAST','FATE','FB','FBHS','FCFS','FCN','FCX','FDS','FDX','FFIV','FGEN','FICO','FIS','FISV','FITB','FIVE',
		'FIVN','FIZZ','FL','FLIR','FLOW','FLR','FLS','FLT','FLWS','FMC','FND','FNKO','FOCS','FOE','FORM','FORR','FOX','FOXA','FOXF',
		'FRHC','FRPH','FRPT','FSCT','FSLR','FSLY','FTI','FTNT','FTV','FUL','GBCI','GBT','GBX','GCO','GCP','GD','GDDY','GDOT','GE',
		'GEF','GGG','GH','GHC','GILD','GIS','GKOS','GL','GLW','GM','GMED','GMS','GNL','GNRC','GOOG','GOOGL','GOSS','GPC','GPI',
		'GPN','GPS','GRA','GRMN','GRUB','GS','GSH','GSKY','GT','GTHX','GTLS','GTN','GTX','GVA','GWRE','GWW','H','HA','HAE',
		'HAIN','HAL','HALO','HAS','HBAN','HBI','HCA','HCCI','HCSG','HD','HDS','HEAR','HEI','HES','HFC','HGV','HHC','HHR',
		'HIBB','HIG','HII','HLT','HNP','HOG','HOLX','HON','HP','HPE','HPQ','HQY','HRB','HRC','HRL','HRTX','HSC','HSIC',
		'HSKA','HST','HSY','HTHT','HUBB','HUBG','HUBS','HUM','HURN','HWM','HXL','IAC','IART','IBM','IBN','IBP','ICE','ICUI',
		'IDCC','IDXX','IEX','IFF','IIVI','ILMN','IMMU','INCY','INFO','INGN','INGR','INSP','INTC','INTU','IONS','IOVA','IP',
		'IPAR','IPG','IPGP','IPHI','IQV','IR','IRBT','IRM','IRTC','ISRG','IT','ITGR','ITRI','ITT','ITW','IVZ','J','JACK',
		'JBHT','JBSS','JCI','JCOM','JD','JEF','JELD','JKHY','JLL','JNJ','JNPR','JOBS','JOUT','JPM','JWN','K','KALU','KCEL',
		'KDP','KEP','KEX','KEY','KEYS','KFY','KHC','KIM','KLAC','KMB','KMI','KMPR','KMT','KMX','KNX','KO','KR','KRG','KSU',
		'KTB','KZTK','L','LAD','LANC','LASR','LB','LDOS','LEA','LECO','LEG','LEGH','LEN','LEVI','LFC','LFUS','LGIH','LGND',
		'LH','LHCG','LHX','LII','LIN','LITE','LKQ','LLY','LM','LMT','LNT','LNTH','LOGM','LOPE','LOW','LPL','LPLA','LPSN',
		'LRCX','LRN','LSTR','LTHM','LULU','LUV','LVS','LW','LYB','LYFT','LYV','M','MA','MAA','MAC','MAN','MANH','MANT','MANU',
		'MAR','MAS','MASI','MAT','MATX','MBT','MBUU','MC','MCD','MCHP','MCK','MCO','MCRI','MD','MDB','MDGL','MDLZ','MDRX',
		'MDT','MED','MEDP','MEI','MELI','MET','MFGP','MGLN','MGY','MHK','MHO','MIDD','MINI','MKC','MKL','MKSI','MKTX','MLCO',
		'MLHR','MLM','MMC','MMI','MMM','MMS','MMSI','MNK','MNRO','MNST','MO','MOMO','MORN','MOS','MOV','MPC','MPWR','MRC',
		'MRK','MRNA','MRO','MRTX','MS','MSA','MSCI','MSFT','MSGE','MSGN','MSGS','MSI','MSM','MSTR','MTB','MTCH','MTD','MTG',
		'MTH','MTN','MTOR','MTRN','MTSC','MU','MUR','MUSA','MXIM','MXL','MYGN','MYL','MYOK','MYRG','NATI','NAVI','NBIX',
		'NBL','NCR','NDAQ','NDSN','NEE','NEM','NEO','NEOG','NEU','NEWR','NFLX','NGVT','NJR','NKE','NKTR','NLOK','NLSN','NMIH',
		'NOC','NOK','NOV','NOW','NRG','NSC','NSIT','NSP','NTAP','NTCO','NTCT','NTES','NTGR','NTLA','NTNX','NTRS','NTUS','NUE',
		'NUS','NUVA','NVDA','NVEE','NVR','NVRO','NVTA','NWL','NWS','NWSA','NXST','NYT','O','OC','ODFL','OFIX','OI','OII','OIS',
		'OKE','OKTA','OLED','OLLI','OMC','OMCL','ON','ONTO','ORCL','ORLY','OSIS','OSK','OSUR','OTIS','OVV','OXY','PAGS','PANW',
		'PATK','PAYC','PB','PBCT','PBF','PBH','PBI','PCAR','PCG','PCRX','PCTY','PD','PDCO','PEAK','PEG','PEGA','PEN','PEP','PETQ',
		'PFE','PFG','PFGC','PFPT','PG','PGR','PGTI','PH','PHM','PII','PINC','PINS','PKG','PKI','PLAN','PLAY','PLCE','PLD','PLNT',
		'PLUS','PLXS','PM','PNC','PNFP','PNTG','PODD','POL','POOL','POST','POWI','PPC','PPG','PPL','PRAA','PRAH','PRFT','PRGS',
		'PRI','PRLB','PRSC','PRSP','PRU','PS','PSA','PSN','PSTG','PSX','PTC','PTR','PUMP','PVH','PWR','PXD','PYPL','PZZA','QADA',
		'QCOM','QDEL','QLYS','QNST','QRTEA','QRVO','QTWO','QUOT','R','RACE','RAMP','RARE','RAVN','RBC','RCL','RDFN','RDS.A','RDY',
		'RE','REG','REGI','REGN','RETA','REX','REZI','RF','RGA','RGEN','RGLD','RGNX','RH','RHI','RIG','RJF','RL','RMD','RNG',
		'ROCK','ROG','ROK','ROKU','ROL','ROLL','ROP','ROST','RP','RPD','RPM','RRBI','RRC','RRGB','RS','RSG','RTX','RVLV','RXN',
		'RYN','RYTM','SAFM','SAGE','SAIA','SAIC','SAIL','SAM','SAVE','SBCF','SBGI','SBH','SBRA','SBUX','SCCO','SCHW','SCSC','SEDG',
		'SEE','SEIC','SERV','SF','SFIX','SFM','SGEN','SHAK','SHEN','SHI','SHW','SIG','SIGI','SINA','SITE','SIVB','SJM','SKM','SKX',
		'SLAB','SLB','SLG','SMAR','SMG','SMPL','SMTC','SNA','SNAP','SNBR','SNPS','SNX','SNY','SO','SOHU','SON','SONO','SP','SPB@US',
		'SPG','SPGI','SPLK','SPR','SPSC','SQ','SRC','SRCL','SRDX','SRE','SRI','SRPT','SSD','SSNC','SSTK','STAA','STLD','STRA','STT',
		'STX','STZ','SUPN','SWAV','SWBI','SWCH','SWI','SWK','SWKS','SWN','SXI','SXT','SYF','SYK','SYKE','SYNA','SYNH','SYY','T',
		'TAK','TAL','TAP','TCBI','TCMD','TCRR','TCS','TCX','TDC','TDG','TDOC','TDS','TDY','TECD','TECH','TEL','TENB','TER','TFC',
		'TFX','TGNA','TGT','THG','THO','THRM','THS','TIF','TJX','TKR','TMHC','TMO','TMUS','TNDM','TNET','TOL','TOT','TPH','TPIC',
		'TPR','TPX','TREE','TREX','TRHC','TRIP','TRMB','TROW','TRU','TRUP','TRV','TSCO','TSLA','TSM','TSN','TT','TTC','TTD','TTEK',
		'TTM','TTMI','TTWO','TWLO','TWNK','TWOU','TWTR','TXN','TXRH','TXT','TYL','UA','UAA','UAL','UBER','UCTT','UDR','UFPI','UFS',
		'UHS','UI','ULTA','UMBF','UNF','UNH','UNM','UNP','UNVR','UPS','UPWK','URBN','URI','USB','USFD','USM','USNA','UTHR','V',
		'VALE','VAR','VC','VCEL','VCRA','VCYT','VEEV','VEON','VFC','VG','VIAC','VICR','VIPS','VLO','VMC','VMI','VMW','VNDA',
		'VNE','VNO','VPG','VREX','VRNS','VRNT','VRSK','VRSN','VRTU','VRTV','VRTX','VTR','VZ','W','WAB','WAL','WAT','WB','WBA',
		'WCC','WDAY','WDC','WDFC','WEC','WELL','WERN','WEX','WFC','WGO','WH','WHD','WHR','WING','WK','WLK','WLTW','WM','WMB',
		'WMT','WOR','WRB','WRK','WRLD','WSM','WSO','WST','WTFC','WTS','WTTR','WU','WWD','WWE','WWW','WY','WYND','WYNN','XEC',
		'XEL','XLNX','XLRN','XNCR','XOM','XPO','XRAY','XRX','XYL','Y','YELP','YETI','YEXT','YUM','YUMC','YY','Z','ZBH','ZBRA',
		'ZEN','ZG','ZGNX','ZION','ZM','ZNH','ZS','ZTS','ZUMZ','ZUO','ZYNE','ZYXI']

def request_page(url: str, ticker: str = "AAPL"):
    """requesting html page from finviz using default AAPL """
    try:
        payload = {"t": ticker}
        req = requests.get(url, params=payload)
        data = bs(req.content, 'html.parser')
        if req.status_code == 200:
            data = bs(req.content, 'lxml')
            return data
        else:
            print ("couldn't get page, status code: ",req.status_code) 
            return
    except Exception as e:
        print (e)
        return
        

def find_data(company: str, multiplicators: list):
    """is looking for company's (like 'FB'/'AAPL') parametrs like ROE, P/E in html page """
    data = request_page(URL, company)
    companys_multi = {}
    try:
        print (data.title.text)
        for param in multiplicators:
            result = data.find(text = param)
            multi = result.next_element.text
            companys_multi[result] = multi
        print (companys_multi)
        return companys_multi
    except Exception as e:
        print (e)


def get_urls(*params: str, num_of_links: int = 1) -> list:
    """because of finviz's limitations, one url can not be used
    Thats why ulrs are splitted if urls length is more than 2900
    fa_pe_u50      p/e<50
    fa_pe_o10      p/e>10

    fa_roe_u00   roe<00
    fa_roe_o15   roe>15

    fa_debteq_u15   debt/equity<15
    fa_debteq_o10   debt/equity>10

    fa_roi_o20    roi>20
    fa_roi_u10    roi<10

    fa_ps_u5    p/s<5
    fa_ps_o1    P/S>1"""
    urls = []
    try:
        for i in range(num_of_links):
            url = "https://finviz.com/screener.ashx?v=111"
            codes = ','.join(rts_codes[len(rts_codes)*(num_of_links - i - 1)//num_of_links:(len(rts_codes)*(num_of_links - i)//num_of_links)])
            payload = {"FT": 2,"f": params,"t": codes}
            req = requests.get(url, params=payload)
            if len(req.url) > 2900:
                urls = []
                num_of_links += 1
                urls = get_urls(*params, num_of_links=num_of_links)
            else:
                urls.append(req.url)
        return (urls)
    except Exception as e:
        print (e)
        return None
    

def get_balance(ticker):
    all_balance = yafin(ticker)
    print (json.dumps(all_balance.get_financial_stmts("annual", "income"), sort_keys=True, indent=4))
    print ((all_balance.get_operating_income()))

def get_mcap(ticker):
    market_cap = yf.Ticker(ticker)
#    print ((json.dumps(market_cap.get_market_cap("2017-08-08","2018-08-08"), sort_keys=True, indent=4)))
    print (dir(market_cap.history))

def get_macro(company, docs):
    req = requests.get("https://www.macrotrends.net/stocks/charts/"+company+"/"+docs)
    data = bs(req.content, 'html.parser')
    result = data.find_all('script')
    var = (re.findall(r"^ var originalData = (.+);", str(result), re.M))
    listik = json.loads(var[0])
    last_dict = {}
    for dictionary in listik:
        del dictionary["popup_icon"]
        changed_field_name = re.findall(r">(.+)</a>", dictionary["field_name"])
        dictionary["field_name"] = changed_field_name
        if dictionary.get("field_name") != []:
            for key, value in sorted(dictionary.items()):
                if key != "field_name":
                    if not last_dict.get(key):
                        last_dict[key] = {}
                        field_name = dictionary["field_name"][0]
                        last_dict[key][field_name] = value.replace(".",",")
                    else:
                        field_name = dictionary["field_name"][0]
                        last_dict[key][field_name] = value.replace(".",",")
    return last_dict

def write_to_exel(our_dict, sheet):
    try:
        book = load_workbook(file_exel)
    except FileNotFoundError:
        book = Workbook()
        book.remove(book['Sheet'])
    finally:
        ws = book.create_sheet(sheet)
        i, j = 2, 2
        formula = "=SUM((C2-B2)/B2)"
        ws['C3'] = formula
        for key in sorted(our_dict.keys()):
            ws.cell(row = 1, column = j).value = key
            for key1, value1 in sorted(our_dict[key].items()):
                if ws.cell(row=i, column=1).value != key1:
                    ws.cell(row=i, column=1).value = key1
                    ws.cell(row=i, column=j).value = value1
                    ws.cell(row=(i+1), column=j).value = Translator(formula, origin="C3").translate_formula(ws.cell(row=(i+1), column=j).coordinate)
                    i += 2
                else:
                    ws.cell(row=i, column=j).value = value1
                    ws.cell(row=(i+1), column=j).value = Translator(formula, origin="C3").translate_formula(ws.cell(row=(i+1), column=j).coordinate)
                    i += 2
            i = 2
            j += 1
        book.save(filename = file_exel)
if __name__ == "__main__":
    try: 
#        print (get_urls("fa_pe_u20"))
        my_dict1 = get_macro(sys.argv[1], "income-statement")
        my_dict2 = get_macro(sys.argv[1], "balance-sheet")
        file_exel = sys.argv[1].replace("/","") + ".xlsx"
        write_to_exel(my_dict1, "income-statement")
        write_to_exel(my_dict2, "balance-sheet")
        write_to_exel(my_dict1, sys.argv[1].replace("/"," ") + " income")
        write_to_exel(my_dict2, sys.argv[1].replace("/"," ") + " balance")
#        get_mcap("AAPL")
    except IndexError as e:
        print ("ERROR! Usage: python3 script.py company_name")
#    get_balance("AAPL")
#    screener("fa_pe_u15")#    print (get_urls("fa_pe_u15", "fa_debteq_u0.1","fa_roe_o25"))

